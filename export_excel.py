import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from database import cursor
from utils import format_rupiah, header

def export_to_excel():
   wb = Workbook()
   ws = wb.active
   header("EXPORT KE EXCEL")

   try:
        bulan = input("Masukkan bulan : ")
   
        bulan = int(bulan)
   
        if bulan < 1 or bulan > 12:
                print("MASUKKAN BULAN YANG BENAR")
                return
        
        bulan = f"{int(bulan):02d}"

   except ValueError:
        print("MASUKKAN ANGKA YANG BENAR")
        return
   
   tahun = input("Masukkan tahun : ")
   ws.title = f"Rekap data bulan {bulan}-{tahun}"
   
   print("\n \nmencari transaksi bulan ",(bulan)," pada tahun ",(tahun))
   
   cursor.execute("""
                    SELECT jenis, kategori, keterangan, jumlah, tanggal
                    FROM transaksi_keuangan
                    WHERE strftime('%m', tanggal) = ?
                    AND strftime('%Y', tanggal) = ?
                    """,(bulan, tahun))
   
   hasil = cursor.fetchall()

   fill_masuk = PatternFill(
         fill_type = "solid",
         start_color = "C6E0B4",
         end_color = "C6E0B4"
      )
   fill_keluar = PatternFill(
         fill_type = "solid",
         start_color = "F8CECC",
         end_color = "F8CECC"
      )
   
   ws.append([
             "JENIS",
             "KATEGORI",
             "KETERANGAN",
             "JUMLAH",
             "TANGGAL"
          ])

   for row in hasil:
       ws.append([
          row[0],
          row[1],
          row[2],
          format_rupiah(abs(row[3])),
          row[4]
       ])

       if row[0] == "MASUK":
          for cel in ws[ws.max_row]:
             cel.fill = fill_masuk

       else:
         for cel in ws[ws.max_row]:
            cel.fill = fill_keluar

   ws.append([])
   ws.append([
      "TOTAL PEMASUKAN",
        format_rupiah(sum(row[3] for row in hasil if row[0] == "MASUK"))])
   ws.append([
      "TOTAL PENGELUARAN",
        format_rupiah(abs(sum(row[3] for row in hasil if row[0] == "KELUAR")))])
   ws.append([
      "TOTAL SALDO",
        format_rupiah(sum(row[3] for row in hasil))])

   header_fill = PatternFill(
         fill_type = "solid",
         start_color = "1F4E78",
         end_color = "1F4E78"
      )   
   
   header_font = Font(
         name = "Times New Roman",
         color = "FFFFFF",
         bold = True
      )
   
   garis = Side(
         border_style = "thin",
         color = "000000"
      )
   
   border = Border(
         left = garis,
         right = garis,
         top = garis,
         bottom = garis
      )
   
   for cel in ws[1]:
         cel.fill = header_fill
         cel.font = header_font
         cel.border = border
   
   for row in ws.iter_rows():
         for cel in row:
            cel.border = border

   for column_cells in ws.columns:
      panjang = 0
      for cell in column_cells:
         if cell.value is not None:
            panjang = max(panjang, len(str(cell.value)))

      nama_kolom = get_column_letter(column_cells[0].column)
      ws.column_dimensions[nama_kolom].width = panjang + 4

   char_sheet = wb.create_sheet("Data Grafik")
   char_sheet["A1"] = "keterangan"
   char_sheet["B1"] = "jumlah"

   char_sheet["A2"] = "TOTAL PEMASUKAN"
   char_sheet["B2"] = sum(row[3] for row in hasil if row[0] == "MASUK")

   char_sheet["A3"] = "TOTAL PENGELUARAN"
   char_sheet["B3"] = abs(sum(row[3] for row in hasil if row[0] == "KELUAR"))

   pie = PieChart()
   pie_data = Reference(
      char_sheet,
      min_col = 2,
      min_row = 1,
      max_row = 3
   )

   pie_labels = Reference(
      char_sheet,
      min_col = 1,
      min_row = 2,
      max_row = 3
   )
   
   for cel in ws['D']:
      cel.alignment = Alignment(horizontal = 'left')

   cursor.execute("""
      SELECT kategori, 
      abs(SUM(jumlah)) AS total
      FROM transaksi_keuangan
      WHERE jenis = 'KELUAR'
      AND strftime('%m', tanggal) = ?
      AND strftime('%Y', tanggal) = ?
      GROUP BY kategori
   """, (bulan, tahun))

   kategori_chart = cursor.fetchall()

   bar_sheet = wb.create_sheet("Data Grafik")

   bar_sheet.append(["kategori", "jumlah"])
   for row in kategori_chart:
      bar_sheet.append(row)

   bar = BarChart()
   bar_data = Reference(
      bar_sheet,
      min_col = 2,
      min_row = 1,
      max_row = bar_sheet.max_row
   )
   bar_labels = Reference(
      bar_sheet,
      min_col = 1,
      min_row = 2,
      max_row = bar_sheet.max_row
   )
   bar.add_data(bar_data, titles_from_data = True)
   bar.set_categories(bar_labels)

   bar.title = "Pengeluaran per kategori"
   bar.y_axis.title = "Rupiah"
   bar.x_axis.title = "kategori"

   ws.add_chart(bar, "H22")
    
   ws.freeze_panes = "A2"
   ws.auto_filter.ref = f"A1:E{ws.max_row}"

   pie.add_data(pie_data, titles_from_data = True)
   pie.set_categories(pie_labels)
   pie.title = "perbandingan keluar dan masuk"
   ws.add_chart(pie, "H2")
   char_sheet.sheet_state = "hidden"

   folder = "export"
   if not os.path.exists(folder):
       os.makedirs(folder, exist_ok=True)
   nama_file = os.path.join(folder, ws.title + ".xlsx")
   wb.save(nama_file)
   print(f"Data berhasil diekspor \n{nama_file}")

def export(user_id, bulan, tahun):
   wb = Workbook()
   ws = wb.active
   try:
        bulan = int(bulan)
        if  bulan < 1 or bulan > 12:
            return "Masukkan bulan yang benar"
        bulan = f"{int(bulan):02d}"

        tahun = tahun
            
   except ValueError:
       return "Format export salah\nContoh :\nexport 8 2026"

   ws.title = f"{bulan}-{tahun}"

   cursor.execute("""
                       SELECT jenis, kategori, keterangan, jumlah, tanggal
                       FROM transaksi_keuangan
                       WHERE user_id = ?
                         AND strftime('%m', tanggal) = ?
                         AND strftime('%Y', tanggal) = ?
                       """,(user_id, bulan, tahun))
      
   hasil = cursor.fetchall() 
   fill_masuk = PatternFill(
         fill_type = "solid",
         start_color = "C6E0B4",
         end_color = "C6E0B4"
      )
   fill_keluar = PatternFill(
         fill_type = "solid",
         start_color = "F8CECC",
         end_color = "F8CECC"
      )
   
   ws.append([
             "JENIS",
             "KATEGORI",
             "KETERANGAN",
             "JUMLAH",
             "TANGGAL"
          ])

   for row in hasil:
       ws.append([
          row[0],
          row[1],
          row[2],
          format_rupiah(abs(row[3])),
          row[4]
       ])

       if row[0] == "MASUK":
          for cel in ws[ws.max_row]:
             cel.fill = fill_masuk

       else:
         for cel in ws[ws.max_row]:
            cel.fill = fill_keluar

   ws.append([])
   ws.append([
      "TOTAL PEMASUKAN",
        format_rupiah(sum(row[3] for row in hasil if row[0] == "MASUK"))])
   ws.append([
      "TOTAL PENGELUARAN",
        format_rupiah(abs(sum(row[3] for row in hasil if row[0] == "KELUAR")))])
   ws.append([
      "TOTAL SALDO",
        format_rupiah(sum(row[3] for row in hasil))])

   header_fill = PatternFill(
         fill_type = "solid",
         start_color = "1F4E78",
         end_color = "1F4E78"
      )   
   
   header_font = Font(
         name = "Times New Roman",
         color = "FFFFFF",
         bold = True
      )
   
   garis = Side(
         border_style = "thin",
         color = "000000"
      )
   
   border = Border(
         left = garis,
         right = garis,
         top = garis,
         bottom = garis
      )
   
   for cel in ws[1]:
         cel.fill = header_fill
         cel.font = header_font
         cel.border = border
   
   for row in ws.iter_rows():
         for cel in row:
            cel.border = border

   for column_cells in ws.columns:
      panjang = 0
      for cell in column_cells:
         if cell.value is not None:
            panjang = max(panjang, len(str(cell.value)))

      nama_kolom = get_column_letter(column_cells[0].column)
      ws.column_dimensions[nama_kolom].width = panjang + 4

   char_sheet = wb.create_sheet("Data Pie")
   char_sheet["A1"] = "keterangan"
   char_sheet["B1"] = "jumlah"

   char_sheet["A2"] = "TOTAL PEMASUKAN"
   char_sheet["B2"] = sum(row[3] for row in hasil if row[0] == "MASUK")

   char_sheet["A3"] = "TOTAL PENGELUARAN"
   char_sheet["B3"] = abs(sum(row[3] for row in hasil if row[0] == "KELUAR"))

   pie = PieChart()
   pie_data = Reference(
      char_sheet,
      min_col = 2,
      min_row = 1,
      max_row = 3
   )

   pie_labels = Reference(
      char_sheet,
      min_col = 1,
      min_row = 2,
      max_row = 3
   )
   
   for cel in ws['D']:
      cel.alignment = Alignment(horizontal = 'left')

   cursor.execute("""
      SELECT kategori, 
      abs(SUM(jumlah)) AS total
      FROM transaksi_keuangan
      WHERE user_id = ?
      AND jenis = 'KELUAR'
      AND strftime('%m', tanggal) = ?
      AND strftime('%Y', tanggal) = ?
      GROUP BY kategori
   """, (user_id, bulan, tahun))

   kategori_chart = cursor.fetchall()

   bar_sheet = wb.create_sheet("Data Kategori")

   bar_sheet.append(["kategori", "jumlah"])
   for row in kategori_chart:
      bar_sheet.append(row)

   bar = BarChart()
   bar_data = Reference(
      bar_sheet,
      min_col = 2,
      min_row = 1,
      max_row = bar_sheet.max_row
   )
   bar_labels = Reference(
      bar_sheet,
      min_col = 1,
      min_row = 2,
      max_row = bar_sheet.max_row
   )
   bar.add_data(bar_data, titles_from_data = True)
   bar.set_categories(bar_labels)

   bar.title = "Pengeluaran per kategori"
   bar.y_axis.title = "Rupiah"
   bar.x_axis.title = "kategori"

   ws.add_chart(bar, "H22")
    
   ws.freeze_panes = "A2"
   ws.auto_filter.ref = f"A1:E{ws.max_row}"

   pie.add_data(pie_data, titles_from_data = True)
   pie.set_categories(pie_labels)
   pie.title = "perbandingan keluar dan masuk"
   ws.add_chart(pie, "H2")
   char_sheet.sheet_state = "hidden"

   folder = "export"
   if not os.path.exists(folder):
       os.makedirs(folder, exist_ok=True)
   nama_file = os.path.join(folder, ws.title + ".xlsx")
   try:
      wb.save(nama_file)
   except PermissionError:
      nama_file = os.path.join(
         folder,
         f"{ws.title}-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
      )
      wb.save(nama_file)
   return f"Data berhasil diekspor \n {nama_file}"
   